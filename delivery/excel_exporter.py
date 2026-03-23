"""Generate Excel exports for compliance state/country maps."""
from __future__ import annotations
import logging
from datetime import date
from pathlib import Path
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)


def _get_workbook():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return openpyxl, Font, PatternFill, Alignment, Border, Side


def generate_pfas_excel(pfas_state_data: Optional[List[dict]] = None, activity_counts: Optional[Dict[str, int]] = None) -> Path:
    """Generate Excel for PFAS state tracker. Returns path."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_workbook()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PFAS State Tracker"

    # Headers
    headers = ["State", "Status", "Tier", "Key Deadline", "Articles This Week", "Notes"]
    header_fill = PatternFill(start_color="2C2C2C", end_color="2C2C2C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Status colors (light versions for Excel)
    status_colors = {
        "comprehensive": "FFCCCC",  # light red
        "limited": "FFF3CC",        # light amber
        "proposed": "CCE5FF",       # light blue
        "none": "F0F0F0",           # light gray
    }

    # Load PFAS state data from config JSON
    import json
    data_path = Path(__file__).parent.parent / "config" / "pfas_state_data.json"
    with open(data_path) as f:
        data = json.load(f)
    states = data["states"]

    for row_idx, (abbr, state) in enumerate(sorted(states.items(), key=lambda x: x[1].get("name", "")), 2):
        status = state.get("status", "none")
        tier = state.get("tier", "")
        deadline = state.get("key_deadline", state.get("deadline", ""))
        activity = activity_counts.get(abbr, 0) if activity_counts else 0
        notes = state.get("notes", state.get("summary", ""))

        row_data = [state.get("name", abbr), status.title(), tier, deadline, activity, notes]

        fill_color = status_colors.get(status, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col <= 2:
                cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 6))

    # Column widths
    for col, width in enumerate([20, 18, 15, 25, 18, 50], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    output = Path(__file__).parent.parent / "data" / f"pfas_tracker_{date.today().isoformat()}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    logger.info(f"PFAS Excel saved: {output}")
    return output


def generate_epr_excel(activity_counts: Optional[Dict[str, int]] = None) -> Path:
    """Generate Excel for EPR state tracker."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_workbook()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EPR State Tracker"

    headers = ["State", "Program Type", "Status", "Key Deadline", "Articles This Week", "Notes"]
    header_fill = PatternFill(start_color="2C2C2C", end_color="2C2C2C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    from delivery.epr_map_generator import _EPR_STATE_DATA

    status_colors = {
        "comprehensive": "CCE5FF",
        "limited": "D4EDDA",
        "proposed": "FFF3CC",
        "none": "F0F0F0",
    }

    for row_idx, (abbr, state) in enumerate(sorted(_EPR_STATE_DATA.items()), 2):
        status = state.get("status", "none")
        activity = activity_counts.get(abbr, 0) if activity_counts else 0

        # Summarise programs as a comma-joined string for the program_type column
        programs = state.get("programs", [])
        program_type = "; ".join(
            p.split(" — ")[0] if " — " in p else p
            for p in programs
        ) if programs else ""

        row_data = [
            state.get("name", abbr),
            program_type,
            status.title(),
            state.get("deadline", ""),
            activity,
            state.get("summary", ""),
        ]

        fill_color = status_colors.get(status, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col <= 3:
                cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (2, 6)))

    for col, width in enumerate([20, 30, 18, 25, 18, 50], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    output = Path(__file__).parent.parent / "data" / f"epr_tracker_{date.today().isoformat()}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    logger.info(f"EPR Excel saved: {output}")
    return output


def generate_reach_excel(activity_counts: Optional[Dict[str, int]] = None) -> Path:
    """Generate Excel for EU REACH country tracker."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_workbook()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EU REACH Tracker"

    headers = ["Country", "Code", "Status", "Supplier Relevance", "Enforcement Body", "Articles This Week", "Notes"]
    header_fill = PatternFill(start_color="2C2C2C", end_color="2C2C2C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    from delivery.reach_map_generator import _REACH_COUNTRY_DATA

    status_colors = {
        "priority": "C8E6C9",   # green
        "high": "DCEDC8",
        "monitor": "F1F8E9",
        "standard": "F9FBE7",
    }

    for row_idx, (code, country) in enumerate(
        sorted(_REACH_COUNTRY_DATA.items(), key=lambda x: x[1].get("name", "")), 2
    ):
        status = country.get("status", "standard")
        activity = activity_counts.get(code, 0) if activity_counts else 0

        row_data = [
            country.get("name", code),
            code,
            status.title(),
            country.get("supplier_relevance", ""),
            country.get("enforcement_body", ""),
            activity,
            country.get("key_notes", ""),
        ]

        fill_color = status_colors.get(status, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col <= 3:
                cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (5, 7)))

    for col, width in enumerate([22, 8, 15, 18, 40, 18, 50], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    output = Path(__file__).parent.parent / "data" / f"reach_tracker_{date.today().isoformat()}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    logger.info(f"REACH Excel saved: {output}")
    return output
